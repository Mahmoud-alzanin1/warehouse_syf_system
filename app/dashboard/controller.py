from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, request, Response
from flask_login import login_required
from sqlalchemy import func, or_

from app.core.database import db
from app.core.permissions import permission_required
from app.models.warehouse import Warehouse
from app.models.inbound import Inbound
from app.models.outbound import Outbound
from app.models.distribution import Distribution


dashboard_bp = Blueprint("dashboard", __name__)


# =========================
# Helpers
# =========================
def _parse_date(value: str):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _warehouse_filter_expr(model, warehouse: Warehouse, legacy_name_field: str):
    """
    فلتر ذكي:
    - لو السجل عليه warehouse_id -> نفلتر بها
    - وإلا fallback على الاسم القديم (fdp/dist_point)
    """
    return or_(
        getattr(model, "warehouse_id") == warehouse.id,
        getattr(model, legacy_name_field) == warehouse.name,
    )


# =========================
# 1) Dashboard Main Page (ALL COMMODITIES)
# =========================
@dashboard_bp.route("/", methods=["GET"])
@login_required
@permission_required("can_access_dashboard")
def index():
    fdp = (request.args.get("fdp", "") or "").strip()
    governorate = (request.args.get("governorate", "") or "").strip()

    wh_query = Warehouse.query.filter_by(is_active=True)
    if fdp:
        wh_query = wh_query.filter(Warehouse.name.ilike(f"%{fdp}%"))

    warehouses = wh_query.order_by(Warehouse.name.asc()).all()

    rows = []
    total_inbound = 0
    total_outbound = 0
    total_balance = 0

    for wh in warehouses:
        # Inbound (ALL commodities)
        in_q = (
            db.session.query(func.sum(Inbound.net_boxes))
            .filter(_warehouse_filter_expr(Inbound, wh, legacy_name_field="fdp"))
        )
        if governorate:
            in_q = in_q.filter(Inbound.governorate.ilike(f"%{governorate}%"))
        inbound_total = int(in_q.scalar() or 0)

        # Outbound (ALL commodities)
        out_q = (
            db.session.query(func.sum(Outbound.net_boxes))
            .filter(_warehouse_filter_expr(Outbound, wh, legacy_name_field="fdp"))
        )
        if governorate:
            out_q = out_q.filter(Outbound.governorate.ilike(f"%{governorate}%"))
        outbound_total = int(out_q.scalar() or 0)

        balance = inbound_total - outbound_total

        rows.append({
            "warehouse_id": wh.id,
            "fdp": wh.name,
            "governorate": governorate or "-",
            "inbound_total": inbound_total,
            "outbound_total": outbound_total,
            "balance": balance,
        })

        total_inbound += inbound_total
        total_outbound += outbound_total
        total_balance += balance

    chart_labels = [r["fdp"] for r in rows]
    chart_inbound = [r["inbound_total"] for r in rows]
    chart_outbound = [r["outbound_total"] for r in rows]
    chart_balance = [r["balance"] for r in rows]

    filters = {"fdp": fdp, "governorate": governorate}

    return render_template(
        "dashboard.html",
        rows=rows,
        filters=filters,
        total_inbound=total_inbound,
        total_outbound=total_outbound,
        total_balance=total_balance,
        chart_labels=chart_labels,
        chart_inbound=chart_inbound,
        chart_outbound=chart_outbound,
        chart_balance=chart_balance,
    )


@dashboard_bp.route("/export", methods=["GET"])
@login_required
@permission_required("can_access_dashboard")
def export():
    """
    تصدير الداشبورد العام (كل الأصناف) إلى Excel:
    لكل مخزن ولكل صنف:
    Inbound / Distribution / Outbound / Balance (Units + MT)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    fdp = (request.args.get("fdp", "") or "").strip()
    governorate = (request.args.get("governorate", "") or "").strip()
    date_from_str = request.args.get("date_from", "") or ""
    date_to_str = request.args.get("date_to", "") or ""

    d_from = _parse_date(date_from_str)
    d_to = _parse_date(date_to_str)

    commodities = ["WHF", "Date Bars", "HEB", "Parcels"]

    wh_query = Warehouse.query.filter_by(is_active=True)
    if fdp:
        wh_query = wh_query.filter(Warehouse.name.ilike(f"%{fdp}%"))
    warehouses = wh_query.order_by(Warehouse.name.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    headers = [
        "Warehouse",
        "Commodity",
        "Inbound Units",
        "Distributed Units",
        "Outbound Units",
        "Balance Units",
        "Inbound MT",
        "Distributed MT",
        "Outbound MT",
        "Balance MT",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for wh in warehouses:
        # inbound grouped
        in_q = db.session.query(
            Inbound.commodity.label("commodity"),
            func.sum(Inbound.net_boxes).label("sum_units"),
            func.sum(Inbound.qty_mt).label("sum_mt"),
        ).filter(_warehouse_filter_expr(Inbound, wh, legacy_name_field="fdp"))

        # outbound grouped
        out_q = db.session.query(
            Outbound.commodity.label("commodity"),
            func.sum(Outbound.net_boxes).label("sum_units"),
            func.sum(Outbound.qty_mt).label("sum_mt"),
        ).filter(_warehouse_filter_expr(Outbound, wh, legacy_name_field="fdp"))

        if governorate:
            in_q = in_q.filter(Inbound.governorate.ilike(f"%{governorate}%"))
            out_q = out_q.filter(Outbound.governorate.ilike(f"%{governorate}%"))

        if d_from:
            in_q = in_q.filter(Inbound.date >= d_from)
            out_q = out_q.filter(Outbound.date >= d_from)
        if d_to:
            in_q = in_q.filter(Inbound.date <= d_to)
            out_q = out_q.filter(Outbound.date <= d_to)

        inbound_rows = in_q.group_by(Inbound.commodity).all()
        outbound_rows = out_q.group_by(Outbound.commodity).all()

        inbound_units = {r.commodity: int(r.sum_units or 0) for r in inbound_rows}
        inbound_mt = {r.commodity: float(r.sum_mt or 0.0) for r in inbound_rows}
        outbound_units = {r.commodity: int(r.sum_units or 0) for r in outbound_rows}
        outbound_mt = {r.commodity: float(r.sum_mt or 0.0) for r in outbound_rows}

        # distribution
        dist_units = {c: 0 for c in commodities}
        dist_mt = {c: 0.0 for c in commodities}

        dq = Distribution.query.filter(_warehouse_filter_expr(Distribution, wh, legacy_name_field="dist_point"))
        if d_from:
            dq = dq.filter(Distribution.dist_date >= d_from)
        if d_to:
            dq = dq.filter(Distribution.dist_date <= d_to)

        for d in dq.all():
            dist_units["WHF"] += d.whf_bags or 0
            dist_units["Date Bars"] += d.date_bars or 0
            dist_units["HEB"] += d.heb_piece or 0
            dist_units["Parcels"] += d.nr_parcels or 0

            dist_mt["WHF"] += d.whf_mt or 0.0
            dist_mt["Date Bars"] += d.date_bars_mt or 0.0
            dist_mt["HEB"] += d.heb_mt or 0.0
            dist_mt["Parcels"] += d.parcels_mt or 0.0

        for c in commodities:
            in_u = inbound_units.get(c, 0)
            dis_u = dist_units.get(c, 0)
            out_u = outbound_units.get(c, 0)
            bal_u = in_u - dis_u - out_u

            in_m = inbound_mt.get(c, 0.0)
            dis_m = dist_mt.get(c, 0.0)
            out_m = outbound_mt.get(c, 0.0)
            bal_m = in_m - dis_m - out_m

            ws.append([
                wh.name,
                c,
                in_u,
                dis_u,
                out_u,
                bal_u,
                round(in_m, 3),
                round(dis_m, 3),
                round(out_m, 3),
                round(bal_m, 3),
            ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# =========================
# 2) Warehouse Dashboard Page (FULL)
# =========================
@dashboard_bp.route("/warehouse/<int:warehouse_id>", methods=["GET"])
@login_required
@permission_required("can_access_dashboard")
def view_warehouse(warehouse_id: int):
    """
    داشبورد مخزن واحد:
    - Inbound + Distribution + Outbound
    - Balance
    - Damage Summary
    - فلترة بالتاريخ
    """

    wh = Warehouse.query.get_or_404(warehouse_id)
    wh_name = wh.name

    date_from_str = request.args.get("date_from", "") or ""
    date_to_str = request.args.get("date_to", "") or ""

    d_from = _parse_date(date_from_str)
    d_to = _parse_date(date_to_str)

    commodities = ["WHF", "Date Bars", "HEB", "Parcels"]

    # ============ 1) Inbound per commodity ============
    inbound_query = (
        db.session.query(
            Inbound.commodity.label("commodity"),
            func.sum(Inbound.net_boxes).label("sum_units"),
            func.sum(Inbound.qty_mt).label("sum_mt"),
            func.sum(Inbound.damage_count).label("sum_damage"),
        )
        .filter(_warehouse_filter_expr(Inbound, wh, legacy_name_field="fdp"))
    )
    if d_from:
        inbound_query = inbound_query.filter(Inbound.date >= d_from)
    if d_to:
        inbound_query = inbound_query.filter(Inbound.date <= d_to)

    inbound_rows = inbound_query.group_by(Inbound.commodity).all()

    inbound_units = {}
    inbound_mt = {}
    damage_inbound = {}
    for r in inbound_rows:
        inbound_units[r.commodity] = int(r.sum_units or 0)
        inbound_mt[r.commodity] = float(r.sum_mt or 0.0)
        damage_inbound[r.commodity] = int(r.sum_damage or 0)

    # ============ 2) Distribution ============
    dist_query = Distribution.query.filter(_warehouse_filter_expr(Distribution, wh, legacy_name_field="dist_point"))
    if d_from:
        dist_query = dist_query.filter(Distribution.dist_date >= d_from)
    if d_to:
        dist_query = dist_query.filter(Distribution.dist_date <= d_to)

    dist_rows = dist_query.all()

    distributed_units = {c: 0 for c in commodities}
    distributed_mt = {c: 0.0 for c in commodities}
    damage_distribution = {c: 0 for c in commodities}

    for d in dist_rows:
        distributed_units["WHF"] += d.whf_bags or 0
        distributed_units["Date Bars"] += d.date_bars or 0
        distributed_units["HEB"] += d.heb_piece or 0
        distributed_units["Parcels"] += d.nr_parcels or 0

        distributed_mt["WHF"] += d.whf_mt or 0.0
        distributed_mt["Date Bars"] += d.date_bars_mt or 0.0
        distributed_mt["HEB"] += d.heb_mt or 0.0
        distributed_mt["Parcels"] += d.parcels_mt or 0.0

        damage_distribution["WHF"] += d.whf_damage_units or 0
        damage_distribution["Date Bars"] += d.date_bars_damage_units or 0
        damage_distribution["HEB"] += d.heb_damage_units or 0
        damage_distribution["Parcels"] += d.parcels_damage_units or 0

    # ============ 3) Outbound per commodity ============
    outbound_query = (
        db.session.query(
            Outbound.commodity.label("commodity"),
            func.sum(Outbound.net_boxes).label("sum_units"),
            func.sum(Outbound.qty_mt).label("sum_mt"),
            func.sum(Outbound.damage_count).label("sum_damage"),
        )
        .filter(_warehouse_filter_expr(Outbound, wh, legacy_name_field="fdp"))
    )
    if d_from:
        outbound_query = outbound_query.filter(Outbound.date >= d_from)
    if d_to:
        outbound_query = outbound_query.filter(Outbound.date <= d_to)

    outbound_rows = outbound_query.group_by(Outbound.commodity).all()

    outbound_units = {}
    outbound_mt = {}
    damage_outbound = {}
    for r in outbound_rows:
        outbound_units[r.commodity] = int(r.sum_units or 0)
        outbound_mt[r.commodity] = float(r.sum_mt or 0.0)
        damage_outbound[r.commodity] = int(r.sum_damage or 0)

    # ============ 4) Merge stats ============
    all_commodities = sorted(set(list(inbound_units.keys()) + commodities + list(outbound_units.keys())))

    stats_list = []
    balance_units = {}
    balance_mt = {}
    damage_stats = []

    for c in all_commodities:
        in_units = inbound_units.get(c, 0)
        dist_units = distributed_units.get(c, 0)
        out_units = outbound_units.get(c, 0)

        in_mt_val = inbound_mt.get(c, 0.0)
        dist_mt_val = distributed_mt.get(c, 0.0)
        out_mt_val = outbound_mt.get(c, 0.0)

        dmg_in = damage_inbound.get(c, 0)
        dmg_dist = damage_distribution.get(c, 0)
        dmg_out = damage_outbound.get(c, 0)
        dmg_balance = dmg_in - (dmg_dist + dmg_out)

        bal_units = in_units - dist_units - out_units
        bal_mt_val = in_mt_val - dist_mt_val - out_mt_val

        balance_units[c] = bal_units
        balance_mt[c] = bal_mt_val

        stats_list.append(
            {"commodity": c, "inbound": in_units, "distributed": dist_units, "outbound": out_units, "balance": bal_units}
        )

        damage_stats.append(
            {
                "commodity": c,
                "damage_inbound": dmg_in,
                "damage_distribution": dmg_dist,
                "damage_outbound": dmg_out,
                "damage_balance": dmg_balance,
            }
        )

    # ============ 5) Chart data ============
    labels = commodities
    inbound_series = [inbound_units.get(c, 0) for c in commodities]
    dist_series = [distributed_units.get(c, 0) for c in commodities]
    balance_series = [balance_units.get(c, 0) for c in commodities]

    filters = {"date_from": date_from_str, "date_to": date_to_str}

    return render_template(
        "warehouse_dashboard.html",
        warehouse_id=warehouse_id,
        wh_name=wh_name,
        filters=filters,
        stats_list=stats_list,
        inbound_mt=inbound_mt,
        distributed_mt=distributed_mt,
        balance_mt=balance_mt,
        damage_stats=damage_stats,
        labels=labels,
        inbound_series=inbound_series,
        dist_series=dist_series,
        balance_series=balance_series,
    )


@dashboard_bp.route("/warehouse/<int:warehouse_id>/export", methods=["GET"])
@login_required
@permission_required("can_access_dashboard")
def export_warehouse(warehouse_id: int):
    """
    تصدير داشبورد مخزن واحد إلى Excel:
    - Inbound / Distribution / Outbound / Balance (Units + MT)
    - Damage summary
    - فلترة بالتاريخ
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wh = Warehouse.query.get_or_404(warehouse_id)

    date_from_str = request.args.get("date_from", "") or ""
    date_to_str = request.args.get("date_to", "") or ""
    d_from = _parse_date(date_from_str)
    d_to = _parse_date(date_to_str)

    # Inbound
    inbound_query = (
        db.session.query(
            Inbound.commodity.label("commodity"),
            func.sum(Inbound.net_boxes).label("sum_units"),
            func.sum(Inbound.qty_mt).label("sum_mt"),
            func.sum(Inbound.damage_count).label("sum_damage"),
        )
        .filter(_warehouse_filter_expr(Inbound, wh, legacy_name_field="fdp"))
    )
    if d_from:
        inbound_query = inbound_query.filter(Inbound.date >= d_from)
    if d_to:
        inbound_query = inbound_query.filter(Inbound.date <= d_to)

    inbound_rows = inbound_query.group_by(Inbound.commodity).all()
    inbound_units = {r.commodity: int(r.sum_units or 0) for r in inbound_rows}
    inbound_mt = {r.commodity: float(r.sum_mt or 0.0) for r in inbound_rows}
    damage_inbound = {r.commodity: int(r.sum_damage or 0) for r in inbound_rows}

    # Distribution
    dist_query = Distribution.query.filter(_warehouse_filter_expr(Distribution, wh, legacy_name_field="dist_point"))
    if d_from:
        dist_query = dist_query.filter(Distribution.dist_date >= d_from)
    if d_to:
        dist_query = dist_query.filter(Distribution.dist_date <= d_to)

    dist_rows = dist_query.all()

    commodities = ["WHF", "Date Bars", "HEB", "Parcels"]
    distributed_units = {c: 0 for c in commodities}
    distributed_mt = {c: 0.0 for c in commodities}
    damage_distribution = {c: 0 for c in commodities}

    for d in dist_rows:
        distributed_units["WHF"] += d.whf_bags or 0
        distributed_units["Date Bars"] += d.date_bars or 0
        distributed_units["HEB"] += d.heb_piece or 0
        distributed_units["Parcels"] += d.nr_parcels or 0

        distributed_mt["WHF"] += d.whf_mt or 0.0
        distributed_mt["Date Bars"] += d.date_bars_mt or 0.0
        distributed_mt["HEB"] += d.heb_mt or 0.0
        distributed_mt["Parcels"] += d.parcels_mt or 0.0

        damage_distribution["WHF"] += d.whf_damage_units or 0
        damage_distribution["Date Bars"] += d.date_bars_damage_units or 0
        damage_distribution["HEB"] += d.heb_damage_units or 0
        damage_distribution["Parcels"] += d.parcels_damage_units or 0

    # Outbound
    outbound_query = (
        db.session.query(
            Outbound.commodity.label("commodity"),
            func.sum(Outbound.net_boxes).label("sum_units"),
            func.sum(Outbound.qty_mt).label("sum_mt"),
            func.sum(Outbound.damage_count).label("sum_damage"),
        )
        .filter(_warehouse_filter_expr(Outbound, wh, legacy_name_field="fdp"))
    )
    if d_from:
        outbound_query = outbound_query.filter(Outbound.date >= d_from)
    if d_to:
        outbound_query = outbound_query.filter(Outbound.date <= d_to)

    outbound_rows = outbound_query.group_by(Outbound.commodity).all()
    outbound_units = {r.commodity: int(r.sum_units or 0) for r in outbound_rows}
    outbound_mt = {r.commodity: float(r.sum_mt or 0.0) for r in outbound_rows}
    damage_outbound = {r.commodity: int(r.sum_damage or 0) for r in outbound_rows}

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"{wh.name} Dashboard"

    ws.append([
        "Commodity",
        "Inbound (Units)",
        "Distributed (Units)",
        "Outbound (Units)",
        "Balance (Units)",
        "Inbound (MT)",
        "Distributed (MT)",
        "Outbound (MT)",
        "Balance (MT)",
        "Damage Inbound",
        "Damage Distribution",
        "Damage Outbound",
        "Damage Balance",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in commodities:
        in_u = inbound_units.get(item, 0)
        dist_u = distributed_units.get(item, 0)
        out_u = outbound_units.get(item, 0)

        in_mt_v = inbound_mt.get(item, 0.0)
        dist_mt_v = distributed_mt.get(item, 0.0)
        out_mt_v = outbound_mt.get(item, 0.0)

        bal_u = in_u - dist_u - out_u
        bal_mt_v = in_mt_v - dist_mt_v - out_mt_v

        dmg_in = damage_inbound.get(item, 0)
        dmg_dist = damage_distribution.get(item, 0)
        dmg_out = damage_outbound.get(item, 0)
        dmg_bal = dmg_in - (dmg_dist + dmg_out)

        ws.append([
            item,
            in_u, dist_u, out_u, bal_u,
            round(in_mt_v, 3),
            round(dist_mt_v, 3),
            round(out_mt_v, 3),
            round(bal_mt_v, 3),
            dmg_in, dmg_dist, dmg_out, dmg_bal
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"warehouse_dashboard_{wh.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
