import os
from datetime import datetime
from io import BytesIO

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    Response,
    current_app,
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from app.core.permissions import permission_required
from app.core.database import db
from app.models.inbound import Inbound

inbound_bp = Blueprint("inbound", __name__, url_prefix="/inbound")


def _apply_filters(query, date_from_str: str, date_to_str: str,
                   employee_name: str, waybill_number: str):
    """فلترة السجلات للأدمن."""
    if date_from_str:
        try:
            d_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            query = query.filter(Inbound.date >= d_from)
        except ValueError:
            pass

    if date_to_str:
        try:
            d_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
            query = query.filter(Inbound.date <= d_to)
        except ValueError:
            pass

    if employee_name:
        query = query.filter(Inbound.inbound_entry_name.ilike(f"%{employee_name.strip()}%"))

    if waybill_number:
        query = query.filter(Inbound.waybill_number.ilike(f"%{waybill_number.strip()}%"))

    return query


def _get_warehouse_id_by_name(name: str):
    """يربط اسم المخزن من الفورم بجدول warehouses ويرجع warehouse_id (fallback)."""
    if not name:
        return None
    try:
        from app.models.warehouse import Warehouse
        wh = Warehouse.query.filter_by(name=name.strip(), is_active=True).first()
        return wh.id if wh else None
    except Exception:
        return None


def _get_warehouse_name_by_id(warehouse_id: int):
    """يرجع اسم المخزن (fdp) بناءً على warehouse_id."""
    if not warehouse_id:
        return None
    try:
        from app.models.warehouse import Warehouse
        wh = Warehouse.query.filter_by(id=warehouse_id, is_active=True).first()
        return wh.name if wh else None
    except Exception:
        return None


def _is_allowed_image(filename: str) -> bool:
    """تحقق بسيط للامتدادات المسموحة."""
    if not filename:
        return False
    _, ext = os.path.splitext(filename)
    ext = (ext or "").lower().lstrip(".")
    allowed = current_app.config.get("ALLOWED_IMAGE_EXTENSIONS", {"png", "jpg", "jpeg", "webp", "jfif", "heic"})

    return ext in allowed


def _safe_waybill_filename(waybill_number: str, ext: str) -> str:
    """
    اسم ملف آمن اعتمادًا على رقم الويبل:
    - نحول أي رموز غريبة إلى _
    - نضيف timestamp لتجنب التعارض
    """
    wb = (waybill_number or "").strip()
    if not wb:
        wb = "waybill"

    # sanitize أقوى من secure_filename لأنه بيحذف أشياء كثير:
    # بنخلي أرقام/حروف/_/-
    cleaned = []
    for ch in wb:
        if ch.isalnum() or ch in ("_", "-", "."):
            cleaned.append(ch)
        else:
            cleaned.append("_")
    base = "".join(cleaned).strip("._-") or "waybill"

    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return f"{base}_{stamp}{ext}"


def _save_waybill_image(file_storage, waybill_number: str) -> str:
    """
    يحفظ صورة الويبل داخل instance/uploads/waybills
    ويرجع مسار للتخزين في DB مثل:
    uploads/waybills/<filename>
    """
    original = secure_filename(file_storage.filename or "")
    if not original:
        raise ValueError("اسم الملف غير صالح.")

    if not _is_allowed_image(original):
        raise ValueError("صيغة الصورة غير مدعومة. استخدم JPG/PNG/WEBP فقط.")

    wb = (waybill_number or "").strip()
    if not wb:
        raise ValueError("رقم الويبل مطلوب لحفظ الصورة باسم رقم الويبل.")

    _, ext = os.path.splitext(original)
    ext = (ext or "").lower()

    save_dir = current_app.config.get("WAYBILLS_UPLOAD_DIR")
    if not save_dir:
        raise RuntimeError("WAYBILLS_UPLOAD_DIR غير مضبوط في config.py")

    os.makedirs(save_dir, exist_ok=True)

    filename = _safe_waybill_filename(wb, ext)
    full_path = os.path.join(save_dir, filename)

    file_storage.save(full_path)

    return f"uploads/waybills/{filename}"


@inbound_bp.route("/", methods=["GET", "POST"])
@login_required
@permission_required("can_access_inbound")
def index():
    """
    - عرض فورم إدخال الويبل
    - حفظ إدخالات جديدة
    - عرض ويبلات مع فلترة + Pagination
    - ✅ يدعم warehouse_id مباشرة + fallback من fdp
    """

    # ---------------------- POST: حفظ إدخال جديد ----------------------
    if request.method == "POST":
        form = request.form

        supervisor_name = (form.get("supervisor_name") or "").strip()

        warehouse_id = form.get("warehouse_id", type=int)
        fdp = (form.get("fdp") or "").strip()

        governorate = (form.get("governorate") or "").strip()
        waybill_number = (form.get("waybill_number") or "").strip()
        commodity = (form.get("commodity") or "").strip()
        si_number = (form.get("si_number") or "").strip()
        date_str = (form.get("date") or "").strip()

        pallets_count = form.get("pallets_count") or "0"
        units_per_pallet = form.get("units_per_pallet") or "0"
        unit_weight_kg = form.get("unit_weight_kg") or "0"
        net_boxes = form.get("net_boxes") or "0"
        activity_type = (form.get("activity_type") or "GFD").strip()
        damage_count = form.get("damage_count") or "0"

        inbound_entry_name = current_user.username

        # تحويل التاريخ (Date)
        try:
            date_value = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            flash("تاريخ غير صالح، الرجاء اختيار تاريخ صحيح.", "danger")
            return redirect(url_for("inbound.index"))

        def to_int(v, default=0):
            try:
                return int(v)
            except (TypeError, ValueError):
                return default

        def to_float(v, default=0.0):
            try:
                return float(v)
            except (TypeError, ValueError):
                return default

        pallets_count = to_int(pallets_count)
        units_per_pallet = to_int(units_per_pallet)
        net_boxes = to_int(net_boxes)
        unit_weight_kg = to_float(unit_weight_kg)
        damage_count = to_int(damage_count)

        if net_boxes == 0 and pallets_count and units_per_pallet:
            net_boxes = pallets_count * units_per_pallet

        qty_mt = None
        if net_boxes and unit_weight_kg:
            qty_mt = (net_boxes * unit_weight_kg) / 1000.0

        # ✅ تحديد المخزن بشكل نهائي (إجباري)
        if warehouse_id:
            resolved_name = _get_warehouse_name_by_id(warehouse_id)
            if not resolved_name:
                flash("المخزن المختار غير صالح أو غير نشط.", "danger")
                return redirect(url_for("inbound.index"))
            fdp = resolved_name
        else:
            warehouse_id = _get_warehouse_id_by_name(fdp)

        if not warehouse_id:
            flash("الرجاء اختيار المخزن بشكل صحيح.", "danger")
            return redirect(url_for("inbound.index"))

        # ✅ حفظ الصورة باسم رقم الويبل
        waybill_image_path = None
        file = request.files.get("waybill_image")
        if file and file.filename:
            try:
                waybill_image_path = _save_waybill_image(file, waybill_number)
            except ValueError as e:
                flash(str(e), "danger")
                return redirect(url_for("inbound.index"))
            except Exception:
                flash("تعذر حفظ الصورة. تأكد من صلاحيات الكتابة داخل instance/uploads.", "danger")
                return redirect(url_for("inbound.index"))

        inbound = Inbound(
            supervisor_name=supervisor_name,
            inbound_entry_name=inbound_entry_name,
            date=date_value,

            fdp=fdp,
            governorate=governorate,
            waybill_number=waybill_number,
            commodity=commodity,
            si_number=si_number,
            pallets_count=pallets_count,
            units_per_pallet=units_per_pallet,
            unit_weight_kg=unit_weight_kg,
            qty_mt=qty_mt,
            net_boxes=net_boxes,
            activity_type=activity_type,
            damage_count=damage_count,
            inbound_datetime=datetime.utcnow(),
            waybill_image=waybill_image_path,
            created_by_user_id=current_user.id,
            movement_type="INBOUND",

            warehouse_id=warehouse_id,
        )

        db.session.add(inbound)
        db.session.commit()

        flash("تم حفظ بيانات الويبل.", "success")
        return redirect(url_for("inbound.index"))

    # ---------------------- GET: عرض السجلات + الفلاتر + Pagination ----------------------
    page = request.args.get("page", 1, type=int)
    per_page = 15

    if getattr(current_user, "is_admin", False):
        date_from_str = request.args.get("date_from", "") or ""
        date_to_str = request.args.get("date_to", "") or ""
        employee_name = request.args.get("employee_name", "") or ""
        waybill_filter = request.args.get("waybill_number", "") or ""

        query = Inbound.query
        query = _apply_filters(query, date_from_str, date_to_str, employee_name, waybill_filter)
    else:
        date_from_str = ""
        date_to_str = ""
        employee_name = ""
        waybill_filter = ""
        query = Inbound.query.filter(Inbound.inbound_entry_name == current_user.username)

    pagination = query.order_by(Inbound.date.desc(), Inbound.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    entries = pagination.items

    filters = {
        "date_from": date_from_str,
        "date_to": date_to_str,
        "employee_name": employee_name,
        "waybill_number": waybill_filter,
    }

    return render_template("inbound.html", entries=entries, filters=filters, pagination=pagination)


@inbound_bp.route("/export", methods=["GET"])
@login_required
@permission_required("can_access_inbound")
def export():
    """تصدير بيانات Inbound إلى Excel .xlsx"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if getattr(current_user, "is_admin", False):
        date_from_str = request.args.get("date_from", "") or ""
        date_to_str = request.args.get("date_to", "") or ""
        employee_name = request.args.get("employee_name", "") or ""
        waybill_filter = request.args.get("waybill_number", "") or ""

        query = Inbound.query
        query = _apply_filters(query, date_from_str, date_to_str, employee_name, waybill_filter)
    else:
        query = Inbound.query.filter(Inbound.inbound_entry_name == current_user.username)

    rows = query.order_by(Inbound.date.asc(), Inbound.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inbound"

    headers = [
        "Date", "FDP", "Governorate", "Waybill", "Commodity", "SI #",
        "No. of Pallets", "Unit/Pallet", "Unit Weight", "QTY. MT",
        "Net No. of Boxes Received", "Activity", "Damage",
        "Supervisor Name", "Inbound Entry Name", "Inbound Datetime",
        "Waybill Image",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    base_url = request.url_root.rstrip("/")
    row_index = 2

    for r in rows:
        ws.append([
            r.date.isoformat() if r.date else "",
            r.fdp or "",
            r.governorate or "",
            r.waybill_number or "",
            r.commodity or "",
            r.si_number or "",
            r.pallets_count or 0,
            r.units_per_pallet or 0,
            r.unit_weight_kg or 0,
            r.qty_mt or 0,
            r.net_boxes or 0,
            r.activity_type or "",
            r.damage_count or 0,
            r.supervisor_name or "",
            r.inbound_entry_name or "",
            r.inbound_datetime.strftime("%Y-%m-%d %H:%M:%S") if r.inbound_datetime else "",
            "",
        ])

        if r.waybill_image:
            img_url = base_url + url_for("files.serve_upload", path=r.waybill_image)
            cell = ws.cell(row=row_index, column=len(headers))
            cell.value = "عرض الويبل"
            cell.hyperlink = img_url
            cell.style = "Hyperlink"

        row_index += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inbound_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@inbound_bp.route("/delete/<int:inbound_id>", methods=["POST"], endpoint="delete")
@login_required
@permission_required("can_access_inbound")
def delete_inbound(inbound_id):
    """حذف سجل Inbound."""
    record = Inbound.query.get_or_404(inbound_id)

    if (not getattr(current_user, "is_admin", False)) and (record.created_by_user_id != current_user.id):
        flash("لا تملك صلاحية حذف هذا السجل.", "danger")
        return redirect(url_for("inbound.index"))

    db.session.delete(record)
    db.session.commit()
    flash("تم حذف الويبل بنجاح.", "success")
    return redirect(url_for("inbound.index"))
