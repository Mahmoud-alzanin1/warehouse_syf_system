import os
from datetime import datetime
from io import BytesIO

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    current_app,
    Response,
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from app.core.database import db
from app.core.permissions import permission_required
from app.models.outbound import Outbound

outbound_bp = Blueprint("outbound", __name__, url_prefix="/outbound")


def _apply_filters(query, date_from_str, date_to_str, fdp, commodity, waybill_number):
    """فلترة سجلات الـ Outbound."""
    if date_from_str:
        try:
            d_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            query = query.filter(Outbound.date >= d_from)
        except ValueError:
            pass

    if date_to_str:
        try:
            d_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
            query = query.filter(Outbound.date <= d_to)
        except ValueError:
            pass

    if fdp:
        query = query.filter(Outbound.fdp == fdp.strip())

    if commodity:
        query = query.filter(Outbound.commodity == commodity.strip())

    if waybill_number:
        query = query.filter(Outbound.waybill_number.ilike(f"%{waybill_number.strip()}%"))

    return query


def _get_warehouse_id_by_name(name: str):
    if not name:
        return None
    try:
        from app.models.warehouse import Warehouse
        wh = Warehouse.query.filter_by(name=name.strip(), is_active=True).first()
        return wh.id if wh else None
    except Exception:
        return None


def _get_warehouse_name_by_id(warehouse_id: int):
    if not warehouse_id:
        return None
    try:
        from app.models.warehouse import Warehouse
        wh = Warehouse.query.filter_by(id=warehouse_id, is_active=True).first()
        return wh.name if wh else None
    except Exception:
        return None


def _is_allowed_image(filename: str) -> bool:
    if not filename:
        return False
    _, ext = os.path.splitext(filename)
    ext = (ext or "").lower().lstrip(".")
    allowed = current_app.config.get("ALLOWED_IMAGE_EXTENSIONS", {"png", "jpg", "jpeg", "webp", "jfif", "heic"})

    return ext in allowed


def _safe_waybill_filename(waybill_number: str, ext: str) -> str:
    wb = (waybill_number or "").strip()
    if not wb:
        wb = "waybill"

    cleaned = []
    for ch in wb:
        if ch.isalnum() or ch in ("_", "-", "."):
            cleaned.append(ch)
        else:
            cleaned.append("_")
    base = "".join(cleaned).strip("._-") or "waybill"

    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return f"{base}_{stamp}{ext}"


def _save_outbound_waybill_image(file_storage, waybill_number: str) -> str:
    original = secure_filename(file_storage.filename or "")
    if not original:
        raise ValueError("اسم الملف غير صالح.")

    if not _is_allowed_image(original):
        raise ValueError("صيغة الصورة غير مدعومة. استخدم JPG/PNG/WEBP فقط.")

    wb = (waybill_number or "").strip()
    if not wb:
        raise ValueError("رقم الويبل مطلوب لحفظ الصورة باسم رقم الويبل.")

    try:
        # تنظيف الاسم
        cleaned = []
        for ch in wb:
            if ch.isalnum() or ch in ("_", "-"):
                cleaned.append(ch)
            else:
                cleaned.append("_")
        wb_clean = "".join(cleaned).strip("_-") or "waybill"

        # تاريخ + وقت
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        public_id = f"{wb_clean}_{timestamp}"

        import cloudinary.uploader
        result = cloudinary.uploader.upload(
            file_storage,
            folder="waybills",
            public_id=public_id,
            overwrite=True,
        )

        return result.get("secure_url")

    except Exception as e:
        raise RuntimeError(f"فشل رفع الصورة إلى Cloudinary: {str(e)}")

@outbound_bp.route("/", methods=["GET", "POST"])
@login_required
@permission_required("can_access_outbound")
def index():
    """
    - إضافة Outbound جديد
    - عرض السجلات مع فلترة + Pagination
    - ✅ يدعم warehouse_id مباشرة + fallback من fdp
    """

    # --------------------- POST: إضافة جديد ---------------------
    if request.method == "POST":
        form = request.form

        date_str = (form.get("date") or "").strip()
        warehouse_id = form.get("warehouse_id", type=int)
        fdp = (form.get("fdp") or "").strip()

        governorate = (form.get("governorate") or "").strip()
        waybill_number = (form.get("waybill_number") or "").strip()
        commodity = (form.get("commodity") or "").strip()
        si_number = (form.get("si_number") or "").strip()

        try:
            pallets_count = int(form.get("pallets_count") or 0)
        except ValueError:
            pallets_count = 0

        try:
            empty_pallets = int(form.get("empty_pallets") or 0)
        except ValueError:
            empty_pallets = 0

        try:
            units_per_pallet = int(form.get("units_per_pallet") or 0)
        except ValueError:
            units_per_pallet = 0

        try:
            unit_weight_kg = float(form.get("unit_weight_kg") or 0)
        except ValueError:
            unit_weight_kg = 0.0

        try:
            damage_count = int(form.get("damage_count") or 0)
        except ValueError:
            damage_count = 0

        try:
            date_val = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            flash("⚠ تاريخ غير صالح، الرجاء اختيار تاريخ صحيح.", "danger")
            return redirect(url_for("outbound.index"))

        net_boxes = pallets_count * units_per_pallet
        qty_mt = (net_boxes * unit_weight_kg) / 1000 if net_boxes and unit_weight_kg else 0

        # ✅ تحديد المخزن بشكل نهائي (إجباري)
        if warehouse_id:
            resolved_name = _get_warehouse_name_by_id(warehouse_id)
            if not resolved_name:
                flash("المخزن المختار غير صالح أو غير نشط.", "danger")
                return redirect(url_for("outbound.index"))
            fdp = resolved_name
        else:
            warehouse_id = _get_warehouse_id_by_name(fdp)

        if not warehouse_id:
            flash("الرجاء اختيار المخزن بشكل صحيح.", "danger")
            return redirect(url_for("outbound.index"))

        # ✅ حفظ الصورة باسم رقم الويبل
        waybill_image_path = None
        file = request.files.get("waybill_image")
        if file and file.filename:
            try:
                waybill_image_path = _save_outbound_waybill_image(file, waybill_number)
            except ValueError as e:
                flash(str(e), "danger")
                return redirect(url_for("outbound.index"))
            except Exception:
                flash("تعذر حفظ الصورة. تأكد من صلاحيات الكتابة داخل instance/uploads.", "danger")
                return redirect(url_for("outbound.index"))

        record = Outbound(
            date=date_val,
            fdp=fdp,
            governorate=governorate,
            waybill_number=waybill_number,
            commodity=commodity,
            si_number=si_number,
            pallets_count=pallets_count,
            empty_pallets=empty_pallets,
            units_per_pallet=units_per_pallet,
            net_boxes=net_boxes,
            unit_weight_kg=unit_weight_kg,
            qty_mt=qty_mt,
            damage_count=damage_count,
            supervisor_name=(form.get("supervisor_name") or "").strip(),
            outbound_entry_name=current_user.username,
            outbound_datetime=datetime.utcnow(),
            waybill_image=waybill_image_path,
            created_by_user_id=current_user.id,
            warehouse_id=warehouse_id,
        )

        db.session.add(record)
        db.session.commit()
        flash("✅ تم حفظ حركة الـ Outbound بنجاح.", "success")
        return redirect(url_for("outbound.index"))

    # --------------------- GET: عرض + فلترة + Pagination ---------------------
    page = request.args.get("page", 1, type=int)
    per_page = 15

    date_from_str = request.args.get("date_from", "") or ""
    date_to_str = request.args.get("date_to", "") or ""
    fdp_filter = request.args.get("fdp", "") or ""
    commodity_filter = request.args.get("commodity", "") or ""
    waybill_filter = request.args.get("waybill_number", "") or ""

    query = Outbound.query
    query = _apply_filters(query, date_from_str, date_to_str, fdp_filter, commodity_filter, waybill_filter)

    pagination = query.order_by(Outbound.date.desc(), Outbound.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    entries = pagination.items

    filters = {
        "date_from": date_from_str,
        "date_to": date_to_str,
        "fdp": fdp_filter,
        "commodity": commodity_filter,
        "waybill_number": waybill_filter,
    }

    return render_template("outbound.html", entries=entries, pagination=pagination, filters=filters)


@outbound_bp.route("/export", methods=["GET"])
@login_required
@permission_required("can_access_outbound")
def export():
    """تصدير حركات الـ Outbound إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    date_from_str = request.args.get("date_from", "") or ""
    date_to_str = request.args.get("date_to", "") or ""
    fdp_filter = request.args.get("fdp", "") or ""
    commodity_filter = request.args.get("commodity", "") or ""
    waybill_filter = request.args.get("waybill_number", "") or ""

    query = Outbound.query
    query = _apply_filters(query, date_from_str, date_to_str, fdp_filter, commodity_filter, waybill_filter)

    rows = query.order_by(Outbound.date.asc(), Outbound.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Outbound"

    headers = [
        "Date", "FDP", "Governorate", "Waybill", "Commodity", "SI #",
        "Pallets Loaded", "Empty Pallets", "Units per Pallet", "Net Boxes",
        "Unit Weight (kg)", "QTY MT", "Damage Count", "Supervisor Name",
        "Outbound Entry Name", "Outbound Datetime", "Waybill Image",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    base_url = request.url_root.rstrip("/")
    row_idx = 2

    for r in rows:
        ws.append([
            r.date.isoformat() if r.date else "",
            r.fdp or "",
            r.governorate or "",
            r.waybill_number or "",
            r.commodity or "",
            r.si_number or "",
            r.pallets_count or 0,
            r.empty_pallets or 0,
            r.units_per_pallet or 0,
            r.net_boxes or 0,
            r.unit_weight_kg or 0,
            r.qty_mt or 0,
            r.damage_count or 0,
            r.supervisor_name or "",
            r.outbound_entry_name or "",
            r.outbound_datetime.strftime("%Y-%m-%d %H:%M:%S") if r.outbound_datetime else "",
            "",
        ])

        if r.waybill_image:
            img_url = base_url + url_for("files.serve_upload", path=r.waybill_image)
            cell = ws.cell(row=row_idx, column=len(headers))
            cell.value = "View Waybill"
            cell.hyperlink = img_url
            cell.style = "Hyperlink"

        row_idx += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"outbound_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@outbound_bp.route("/delete/<int:out_id>", methods=["POST"], endpoint="delete")
@login_required
@permission_required("can_access_outbound")
def delete(out_id):
    rec = Outbound.query.get_or_404(out_id)

    if (not getattr(current_user, "is_admin", False)) and (rec.created_by_user_id != current_user.id):
        flash("لا تملك صلاحية حذف هذا السجل.", "danger")
        return redirect(url_for("outbound.index"))

    db.session.delete(rec)
    db.session.commit()
    flash("✅ تم حذف سجل الـ Outbound", "success")
    return redirect(url_for("outbound.index"))