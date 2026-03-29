import os
from datetime import datetime
from io import BytesIO

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    Response,
    current_app,
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from app.core.database import db
from app.core.permissions import permission_required
from app.models.data_entry import DataEntry
from app.models.warehouse import Warehouse

data_entry_bp = Blueprint("data_entry", __name__)

ALLOWED_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}


def _apply_filters(query, date_from_str: str, date_to_str: str, employee_name: str):
    if date_from_str:
        try:
            d_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            query = query.filter(DataEntry.entry_date >= d_from)
        except ValueError:
            pass

    if date_to_str:
        try:
            d_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
            query = query.filter(DataEntry.entry_date <= d_to)
        except ValueError:
            pass

    if employee_name:
        query = query.filter(DataEntry.data_entry_name.ilike(f"%{employee_name.strip()}%"))

    return query


def _get_warehouse_id_by_name(name: str):
    if not name:
        return None
    try:
        wh = Warehouse.query.filter_by(name=name.strip(), is_active=True).first()
        return wh.id if wh else None
    except Exception:
        return None


def _save_beneficiaries_image(file_storage) -> str:
    original = secure_filename(file_storage.filename or "")
    if not original:
        raise ValueError("اسم الملف غير صالح.")

    _, ext = os.path.splitext(original)
    ext = (ext or "").lower()

    if ext not in ALLOWED_IMAGE_EXTS:
        raise ValueError("⚠ امتداد الصورة غير مسموح. استخدم JPG/PNG/WEBP فقط.")

    save_dir = current_app.config.get("BENEFICIARIES_UPLOAD_DIR")
    if not save_dir:
        raise RuntimeError("BENEFICIARIES_UPLOAD_DIR غير مضبوط في config.py")

    os.makedirs(save_dir, exist_ok=True)
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"beneficiaries_{stamp}{ext}"
    full_path = os.path.join(save_dir, filename)

    file_storage.save(full_path)
    return f"uploads/beneficiaries/{filename}"


@data_entry_bp.route("/", methods=["GET", "POST"])
@login_required
@permission_required("can_access_data_entry")
def index():
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name.asc()).all()

    if request.method == "POST":
        entry_date_str = (request.form.get("entry_date") or "").strip()
        warehouse_name = (request.form.get("warehouse") or "").strip()

        pit_account_name = (request.form.get("pit_account_name") or "").strip()
        parcels_count = request.form.get("parcels_count") or "0"
        beneficiaries_count = request.form.get("beneficiaries_count") or "0"
        distribution_type = (request.form.get("distribution_type") or "").strip()
        data_entry_name = current_user.username
        notes = (request.form.get("notes") or "").strip()

        try:
            entry_date = datetime.strptime(entry_date_str, "%Y-%m-%d").date()
        except ValueError:
            flash("تاريخ غير صالح، الرجاء اختيار تاريخ صحيح.", "danger")
            return redirect(url_for("data_entry.index"))

        def to_int(v, default=0):
            try:
                return int(v)
            except (TypeError, ValueError):
                return default

        parcels_count = to_int(parcels_count)
        beneficiaries_count = to_int(beneficiaries_count)

        warehouse_id = _get_warehouse_id_by_name(warehouse_name)

        if warehouse_name and not warehouse_id:
            flash("⚠ المخزن المختار غير موجود أو غير نشط. اختر مخزن صحيح.", "danger")
            return redirect(url_for("data_entry.index"))

        beneficiaries_image_path = None
        file = request.files.get("beneficiaries_image")
        if file and file.filename:
            try:
                beneficiaries_image_path = _save_beneficiaries_image(file)
            except ValueError as e:
                flash(str(e), "danger")
                return redirect(url_for("data_entry.index"))
            except Exception:
                flash("تعذر حفظ الصورة. تأكد من صلاحيات الكتابة داخل instance/uploads.", "danger")
                return redirect(url_for("data_entry.index"))

        record = DataEntry(
            entry_date=entry_date,
            warehouse=warehouse_name,
            pit_account_name=pit_account_name,
            parcels_count=parcels_count,
            beneficiaries_count=beneficiaries_count,
            distribution_type=distribution_type,
            data_entry_name=data_entry_name,
            notes=notes,
            beneficiaries_image=beneficiaries_image_path,
            created_by_user_id=current_user.id,
            warehouse_id=warehouse_id,
        )

        db.session.add(record)
        db.session.commit()

        flash("✅ تم حفظ بيانات الداتا انتري.", "success")
        return redirect(url_for("data_entry.index"))

    page = request.args.get("page", 1, type=int)
    per_page = 15

    if getattr(current_user, "is_admin", False):
        date_from_str = request.args.get("date_from", "") or ""
        date_to_str = request.args.get("date_to", "") or ""
        employee_name = request.args.get("employee_name", "") or ""

        query = DataEntry.query
        query = _apply_filters(query, date_from_str, date_to_str, employee_name)
    else:
        date_from_str = ""
        date_to_str = ""
        employee_name = ""
        query = DataEntry.query.filter(DataEntry.data_entry_name == current_user.username)

    pagination = query.order_by(DataEntry.entry_date.desc(), DataEntry.id.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    entries = pagination.items

    filters = {
        "date_from": date_from_str,
        "date_to": date_to_str,
        "employee_name": employee_name,
    }

    return render_template(
        "data_entry.html",
        entries=entries,
        filters=filters,
        warehouses=warehouses,
        pagination=pagination,
    )


@data_entry_bp.route("/export", methods=["GET"])
@login_required
@permission_required("can_access_data_entry")
def export():
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if getattr(current_user, "is_admin", False):
        date_from_str = request.args.get("date_from", "") or ""
        date_to_str = request.args.get("date_to", "") or ""
        employee_name = request.args.get("employee_name", "") or ""

        query = DataEntry.query
        query = _apply_filters(query, date_from_str, date_to_str, employee_name)
    else:
        query = DataEntry.query.filter(DataEntry.data_entry_name == current_user.username)

    rows = query.order_by(DataEntry.entry_date.asc(), DataEntry.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Entry"

    headers = [
        "Date",
        "Warehouse",
        "PIT Account",
        "Parcels",
        "Beneficiaries",
        "Distribution Type",
        "Data Entry Name",
        "Notes",
        "Image",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    base_url = request.url_root.rstrip("/")
    row_index = 2

    for r in rows:
        ws.append([
            r.entry_date.isoformat() if r.entry_date else "",
            r.warehouse or "",
            r.pit_account_name or "",
            r.parcels_count or 0,
            r.beneficiaries_count or 0,
            r.distribution_type or "",
            r.data_entry_name or "",
            r.notes or "",
            "",
        ])

        if r.beneficiaries_image:
            img_url = base_url + url_for("files.serve_upload", path=r.beneficiaries_image)
            cell = ws.cell(row=row_index, column=len(headers))
            cell.value = "عرض الصورة"
            cell.hyperlink = img_url
            cell.style = "Hyperlink"

        row_index += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"data_entries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@data_entry_bp.route("/delete/<int:entry_id>", methods=["POST"], endpoint="delete")
@login_required
@permission_required("can_access_data_entry")
def delete_entry(entry_id):
    record = DataEntry.query.get_or_404(entry_id)

    if (not getattr(current_user, "is_admin", False)) and (record.created_by_user_id != current_user.id):
        flash("لا تملك صلاحية حذف هذا السجل.", "danger")
        return redirect(url_for("data_entry.index"))

    db.session.delete(record)
    db.session.commit()
    flash("✅ تم حذف إدخال الداتا انتري.", "success")
    return redirect(url_for("data_entry.index"))
